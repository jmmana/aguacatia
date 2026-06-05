"""
Prepara el dataset Mendeley para entrenamiento con YOLOv8.

Uso:
    python prepare_dataset.py --zip ~/Downloads/"Hass Avocado Ripening Photographic Dataset.zip"

Salida:
    dataset/raw/
        unripe/
        breaking/
        ripe_first/
        ripe_second/
        overripe/
"""

import argparse
import shutil
import zipfile
from pathlib import Path

import openpyxl


# Mapeo de etiquetas del Excel a nombres de carpeta del modelo
LABEL_MAP = {
    "unripe":       "unripe",
    "breaking":     "breaking",
    "ripe":         "ripe_first",       # "Ripe" en el dataset = primer estadio
    "ripe (2)":     "ripe_second",      # variante usada en algunas versiones del Excel
    "ripe_1":       "ripe_first",
    "ripe_2":       "ripe_second",
    "ripe first":   "ripe_first",
    "ripe second":  "ripe_second",
    "overripe":     "overripe",
}

CLASSES = ["unripe", "breaking", "ripe_first", "ripe_second", "overripe"]


def extract_zip(zip_path: Path, dest: Path) -> Path:
    print(f"Extrayendo {zip_path.name}...")
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(dest)
    # El ZIP contiene una carpeta con el nombre del dataset
    extracted = [p for p in dest.iterdir() if p.is_dir()]
    return extracted[0]


def read_labels(xlsx_path: Path) -> dict[str, str]:
    """Lee el Excel y devuelve {nombre_archivo: clase}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Detectar columnas clave
    try:
        img_col = next(i for i, h in enumerate(headers) if "file" in h or "image" in h or "name" in h)
        lbl_col = next(i for i, h in enumerate(headers) if "ripe" in h or "stage" in h or "label" in h or "class" in h)
    except StopIteration:
        # Fallback: asumir columnas 0 y última
        img_col, lbl_col = 0, len(headers) - 1

    labels: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[img_col]:
            continue
        filename = str(row[img_col]).strip()
        raw_label = str(row[lbl_col]).strip().lower() if row[lbl_col] else ""
        mapped = LABEL_MAP.get(raw_label)
        if mapped:
            # Asegurar extensión
            if not filename.endswith(".jpg"):
                filename += ".jpg"
            labels[filename] = mapped

    wb.close()
    return labels


def organize(dataset_dir: Path, images_dir: Path, labels: dict[str, str], out_dir: Path):
    """Copia las imágenes a carpetas por clase."""
    for cls in CLASSES:
        (out_dir / cls).mkdir(parents=True, exist_ok=True)

    found, missing = 0, 0
    for filename, clase in labels.items():
        src = images_dir / filename
        if not src.exists():
            missing += 1
            continue
        dst = out_dir / clase / filename
        shutil.copy2(src, dst)
        found += 1

    print(f"  Imágenes organizadas: {found}")
    if missing:
        print(f"  Imágenes no encontradas en ZIP: {missing}")

    # Resumen por clase
    print("\nDistribución por clase:")
    for cls in CLASSES:
        count = len(list((out_dir / cls).glob("*.jpg")))
        print(f"  {cls:15s}: {count:5d} imágenes")


def main():
    parser = argparse.ArgumentParser(description="Prepara el dataset Mendeley para YOLOv8")
    parser.add_argument("--zip", required=True, help="Ruta al ZIP del dataset Mendeley")
    parser.add_argument("--out", default="raw", help="Carpeta de salida (default: raw)")
    args = parser.parse_args()

    zip_path = Path(args.zip).expanduser().resolve()
    script_dir = Path(__file__).parent
    tmp_dir = script_dir / "_tmp_extract"
    out_dir = script_dir / args.out

    if not zip_path.exists():
        raise FileNotFoundError(f"No se encontró el ZIP: {zip_path}")

    print("=== Preparación del dataset Mendeley ===\n")

    # 1. Extraer ZIP
    tmp_dir.mkdir(exist_ok=True)
    dataset_root = extract_zip(zip_path, tmp_dir)

    # 2. Localizar Excel e imágenes
    xlsx_files = list(dataset_root.rglob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError("No se encontró el archivo Excel de etiquetas en el ZIP")
    xlsx_path = xlsx_files[0]
    print(f"Excel de etiquetas: {xlsx_path.name}")

    images_dir = dataset_root / "Avocado Ripening Dataset"
    if not images_dir.exists():
        # Buscar carpeta de imágenes
        candidates = [d for d in dataset_root.rglob("*") if d.is_dir() and len(list(d.glob("*.jpg"))) > 100]
        images_dir = candidates[0] if candidates else dataset_root

    print(f"Carpeta de imágenes: {images_dir.name}")
    print(f"Total JPGs encontrados: {len(list(images_dir.glob('*.jpg')))}\n")

    # 3. Leer etiquetas del Excel
    print("Leyendo etiquetas del Excel...")
    labels = read_labels(xlsx_path)
    print(f"  Etiquetas leídas: {len(labels)}\n")

    # 4. Organizar por clase
    print(f"Organizando en {out_dir}...")
    organize(dataset_root, images_dir, labels, out_dir)

    # 5. Limpiar temporal
    shutil.rmtree(tmp_dir)

    print(f"\nDataset listo en: {out_dir}")
    print("Siguiente paso: python split_dataset.py")


if __name__ == "__main__":
    main()
